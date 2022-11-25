from datetime import datetime, timedelta
import os.path
from os import remove as os_remove
import base64

from fast_bitrix24 import Bitrix
import openpyxl
from openpyxl.styles import Font

from authentication import authentication


b = Bitrix(authentication('Bitrix'))

users_info = b.get_all('user.get', {'filter': {'ACTIVE': 'true'}})

report_created_time = datetime.now()
report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
report_name = f'Отчет по сделкам {report_name_time}.xlsx'.replace(' ', '_')


def get_deals():
    current_day = datetime.now()
    filter_closedate_end = datetime.strftime(current_day + timedelta(days=31), '%Y-%m-%d')
    filter_closedate_start = datetime.strftime(current_day - timedelta(days=1), '%Y-%m-%d')
    b24_deals = b.get_all('crm.deal.list', {
        'select': [
            'ASSIGNED_BY_ID',
            'STAGE_ID',
            'UF_CRM_1657878818384'
        ],
        'filter': {
            '>CLOSEDATE': filter_closedate_start,
            '<CLOSEDATE': filter_closedate_end,
            'CATEGORY_ID': '1',
            'UF_CRM_1657878818384': [   # Группа
                '859',  # ИТС
                '905',  # Отчетность
            ]

        }})
    return b24_deals


def sort_users():
    global users_info
    departments = {
        'ГО4': [],
        'ГО3': [],
        'ОВ': [],
        'Прочие': [],
    }
    deal_stages = {

    }
    for user in users_info:
        if 29 in user['UF_DEPARTMENT']:
            departments['ГО4'].append(user['ID'])
        elif 27 in user['UF_DEPARTMENT']:
            departments['ГО3'].append(user['ID'])
        elif 225 in user['UF_DEPARTMENT'] or user['ID'] == '109':
            departments['ОВ'].append(user['ID'])
        else:
            if user['ID'] == '109':     # Светлана Ридкобород
                continue
            departments['Прочие'].append(user['ID'])
    return departments


def sort_data_by_group(b24_deals, user_departments, stage_ids, group_name):
    global users_info
    title_date_filter_range = f"{datetime.strftime(datetime.now(), '%d.%m.%Y')} - {datetime.strftime(datetime.now() + timedelta(days=30), '%d.%m.%Y')}"
    group_id = {
        'ИТС': '859',
        'Отчетность': '905',
    }
    result = [
        [f'Состояние сделок, заканчивающихся в период {title_date_filter_range}'],
        ['', '', '', '', 'Услуга активна', 'Счет сформирован', 'Счет отправлен', 'Нет оплаты', 'Ждем решения клиента', 'Услуга завершена', 'Отказ от сопровождения']
    ]
    for department in user_departments:
        department_row = [department, '', '', '']
        for stage in stage_ids:
            filtered_deals_count = len(
                list(
                    filter(
                        lambda x: x['STAGE_ID'] == stage
                                  and x['ASSIGNED_BY_ID'] in user_departments[department]
                                  and x['UF_CRM_1657878818384'] == group_id[group_name],
                        b24_deals
                    )))
            if filtered_deals_count == 0:
                filtered_deals_count = ''
            department_row.append(filtered_deals_count)
        result.append(department_row)

        department_user_rows = []
        for user_id in user_departments[department]:
            user_info = list(filter(lambda x: x['ID'] == user_id, users_info))[0]
            user_name = f"{user_info['LAST_NAME']} {user_info['NAME']}"
            user_row = ['', user_name, '', '']
            for stage in stage_ids:
                filtered_deals_count = len(
                    list(
                        filter(
                            lambda x: x['STAGE_ID'] == stage
                            and x['ASSIGNED_BY_ID'] == user_id
                            and x['UF_CRM_1657878818384'] == group_id[group_name],
                            b24_deals
                        )))
                if filtered_deals_count == 0:
                    filtered_deals_count = ''
                user_row.append(filtered_deals_count)
            if any(user_row[2:]):
                department_user_rows.append(user_row)
        department_user_rows = list(sorted(department_user_rows, key=lambda x: x[1]))
        result += department_user_rows
    return result


def write_data_to_xlsx(data, report_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'ИТС'
    if os.path.exists(report_name):
        workbook = openpyxl.load_workbook(report_name)
        worksheet = workbook.create_sheet('Отчетность')

    for row in data:
        worksheet.append(row)

    bold_cells = ['A', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    for row in range(1, max_rows + 1):
        if worksheet.cell(row, 1).value in ['ГО4', 'ГО3', 'ОВ', 'Прочие']:
            for cell_letter in bold_cells:
                cell_number = cell_letter + str(row)
                worksheet[cell_number].font = Font(bold=True)

    workbook.save(report_name)


def upload_report_to_b24(report_name):
    bitrix_folder_id = '234509'

    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_file = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    os_remove(report_name)
    return upload_file["ID"]


def send_deal_statistic_email():
    stage_ids = [
        'C1:NEW',           # Услуга активна
        'C1:UC_0KJKTY',     # Счет сформирован
        'C1:UC_3J0IH6',     # Счет отправлен клиенту
        'C1:UC_KZSOR2',     # Нет оплаты
        'C1:UC_VQ5HJD',     # Ждём решения клиента
        'C1:WON',           # Услуга завершена
        'C1:LOSE',          # Отказ от сопровождения
    ]
    group_names = ['ИТС', 'Отчетность']
    b24_deals = get_deals()
    user_departments = sort_users()
    for group in group_names:
        data_to_write = sort_data_by_group(b24_deals, user_departments, stage_ids, group)
        write_data_to_xlsx(data_to_write, report_name)
    b24_file_id = upload_report_to_b24(report_name)
    b.call('bizproc.workflow.start', {'TEMPLATE_ID': '1245', 'DOCUMENT_ID': ['lists', 'BizprocDocument', '237723'], 'PARAMETERS': {'file_id': b24_file_id}})


if __name__ == '__main__':
    send_deal_statistic_email()

